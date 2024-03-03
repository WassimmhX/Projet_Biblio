from PyQt5 import QtWidgets
from PyQt5.QtWidgets import *
import xlsxwriter
from datetime import date
import openpyxl
from ajouter_etd import Ui_Dialog
from supprimer_etd import Ui_Dialog1
from modifier_etd import Ui_Dialog2
from ajouter_lv import Ui_Dialog3
from supprimer_lv import Ui_Dialog4
from modifier_lv import Ui_Dialog5
from ajouter_emp import Ui_Dialog6
from supprimer_emp import Ui_Dialog7
from modifier_date_emp import Ui_Dialog8
from modifier_date_ret import Ui_Dialog9
from afficher_etd import Ui_Dialog10
from afficher_lv import Ui_Dialog11
from afficher_emp import Ui_Dialog12
from rech_par_id import Ui_Dialog13
from rech_par_sec import Ui_Dialog14
from rech_par_niv import Ui_Dialog15
from rech_par_cl import Ui_Dialog16
from rech_par_ref import Ui_Dialog17
from rech_par_tit import Ui_Dialog18
from rech_par_aut import Ui_Dialog19
from rech_par_ann import Ui_Dialog20
from rech_par_etd import Ui_Dialog21
from rech_par_liv import Ui_Dialog22
from rech_par_demp import Ui_Dialog23
from rech_par_dret import Ui_Dialog24
from rech_entre_demp import Ui_Dialog25
from rech_entre_dret import Ui_Dialog26
from retour_emp import Ui_Dialog27
workbook1 = openpyxl.load_workbook("etudiant.xlsx")
sheet1 = workbook1.active
dataetd = list(sheet1.values)

workbook2 = openpyxl.load_workbook("livre.xlsx")
sheet2 = workbook2.active
datalv = list(sheet2.values)

workbook3 = openpyxl.load_workbook("emprunt.xlsx")
sheet3 = workbook3.active
dataemp = list(sheet3.values)
def majus(ch):
    l = ch.split()
    for i in range(len(l)):
        l[i] = l[i][0].upper() + l[i][1:]
    ch = " ".join(l)
    return ch


def recherche(ch, data, x):
    for i in data[1:]:
        if i[x] == ch:
            return i


def recherche2(data, x,ch1="", num=""):
    if num == "":
        for i in data[1:]:
            j = 0
            while j < len(i[x])-1 and not (i[x][j].isnumeric()):
                j += 1
            if j == len(i[x])-1:
                ch2 = i[x][:j]
            else:
                ch2 = i[x][:j] + i[x][j + 1:]
            if ch1 == ch2:
                return i
    elif ch1 == "":
        for i in data[1:]:
            if (num in i[x]):
                return i
    for i in data[1:]:
        if (num in i[x]):
            j = 0
            while j < len(i[x]) - 1 and not (i[x][j].isnumeric()):
                j += 1
            if j == len(i[x]) - 1:
                ch2 = i[x][:j]
            else:
                ch2 = i[x][:j] + i[x][j + 1:]
            if ch1 == ch2:
                return i


def recherche3(ch1,x,ch2,y,data):
    for i in range (len(data[1:]),-1,-1):
        if data[i][x] == ch1 and data[i][y]==ch2:
            return data[i]


def enregistrement(l1, data):
    workbook = xlsxwriter.Workbook(data)
    worksheet = workbook.add_worksheet(data[:len(data) - 5])
    w = openpyxl.load_workbook(data)
    sheet = w.active
    if not (type(l1[0]) == tuple):
        l = list(sheet.values)
        l.append(l1)
    else:
        l = l1
    for i in range(len(l)):
        for j in range(0, len(l[i])):
            worksheet.write(i, j, str(l[i][j]))
    workbook.close()


def existe(ch, l1, x):
    for i in l1[1:]:
        if i[x] == ch:
            return True
    return False


def existeCl(ch, l1, x):
    if ch.isnumeric():
        for i in l1[1:]:
            print(i[x])
            if ch in i[x]:
                return True
        return False
    else:
        for i in l1[1:]:
            j = 0
            while j < len(i[x])-1 and not (i[x][j].isnumeric()):
                j += 1
            if j == len(i[x])-1:
                ch2 = i[x][:j]
            else:
                ch2 = i[x][:j] + i[x][j + 1:]
            if ch == ch2:
                return True
        return False

def existeCl1(ch, l1, x):
    if ch.isnumeric():
        if ch in l1[x]:
            return True
        return False
    else:
        j = 0
        while j < len(l1[x]) - 1 and not (l1[x][j].isnumeric()):
            j += 1
        if j == len(l1[x]) - 1:
            ch2 = l1[x][:j]
        else:
            ch2 = l1[x][:j] + l1[x][j + 1:]

        if ch == ch2:
            return True
    return False


def existeIn(ch,l1,x):
    for i in l1[1:]:
        if ch in i[x]:
            return True
    return False


def existeDouble(ch1,x,ch2,y,l1):
    for i in l1[1:]:
        if i[x] == ch1 and i[y]==ch2:
            return True
    return False


def existeDoubleCl(ch1,ch2,y,l1):

    ch1=str(ch1)
    ch2=str(ch2)
    for i in l1[1:]:
        if existeCl1(ch1,i,y) and existeCl1(ch2,i,y):
            return True
    return False


def dateinf(ch1, ch2):
    l1 = ch1.split("/")
    l2 = ch2.split("/")
    if int(l1[2]) < int(l2[2]):
        return True
    elif int(l1[2]) == int(l2[2]):
        if int(l1[1]) < int(l2[1]):
            return True
        elif int(l1[1]) == int(l2[1]):
            if int(l1[0]) <= int(l2[0]):
                return True
    return False


class ajouterEtudiant:
    def __init__(self):
        self.window = QDialog()
        self.ui = Ui_Dialog()
        self.ui.setupUi(self.window)
        self.ui.pushButton.clicked.connect(self.ajouter)

    def ajouter(self):
        global dataetd
        ch = self.ui.lineEdit.text()
        ch2 = self.ui.lineEdit_2.text()
        ch3 = self.ui.lineEdit_3.text()
        ch4 = self.ui.lineEdit_4.text()
        ch5 = self.ui.lineEdit_5.text()
        ch7 = self.ui.lineEdit_7.text()
        ch8 = self.ui.comboBox.currentText()
        ch9 = self.ui.dateEdit.text()
        if ch == "" or ch2 == "" or ch3 == "" or ch4 == "" or ch5 == "" or ch7 == "" or ch8 == "":
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Il y a un champ vide")
        elif not (ch3.isnumeric() and ch7.isnumeric()) or len(ch3) != 8 or len(ch7) != 8:
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "N° d'inscription et N° téléphone "
                                                                             "doivent être 8 chiffres")
        elif ch4[0] == "@" or ("@gmail.com" not in ch4) or not ch4[0:ch4.index("@")].isalnum():
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Mail est incorrect")
        elif not (ch.isalpha() and ch2.isalpha()):
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING",
                                                     "le nom el le prenom ne doivent contenir que des lettres")
        elif not ch5.isalnum():
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING",
                                                     "l'adresse ne doit contenir que des lettres et des nombres")
        else:
            if existe(ch3,dataetd, 2):
                self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cet Etudiant déja existe!")
            else:
                l1 = [majus(ch), majus(ch2), ch3, ch4, ch5, ch7, ch8, ch9]
                dataetd.append(l1)
                enregistrement(l1, "etudiant.xlsx")
                self.box = QtWidgets.QMessageBox.about(self.window, "Succés", "L'ajout est effectué avec succés")


class supprimerEtudiant:
    def __init__(self):
        self.window = QDialog()
        self.ui = Ui_Dialog1()
        self.ui.setupUi(self.window)
        self.ui.pushButton.clicked.connect(self.supprimer)

    def supprimer(self):
        global dataetd
        ch = self.ui.lineEdit.text()
        ch2 = self.ui.comboBox.currentText()
        ch3 = self.ui.comboBox_2.currentText()
        if ch == "" and ch2 == "" and ch3 == "":
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "il faut remplir au moins un champ")
        if not (ch == "" and ch2 == "" and ch3 == ""):
            if ch != "":
                if len(ch) != 8:
                    self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING",
                                                             "N° d'inscription doit être 8 chiffres")
                else:
                    if not (existe(ch, dataetd, 2)):
                        self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cet Etudiant n'existe pas!")
                    else:
                        self.box = QtWidgets.QMessageBox.question(self.window, "question",
                                                                  "Voulez-vous vraiment enregistrer la suppression d'etudiant?")
                        if self.box == QMessageBox.Yes:
                            dataetd.remove(recherche(ch, dataetd, 2))

                            self.box = QtWidgets.QMessageBox.about(self.window, "Succés",
                                                                   "La suppression d'etudiant est effectué "
                                                                   "avec succés")
                        else:
                            self.box = QtWidgets.QMessageBox.about(self.window, "Annuler",
                                                                   "La suppression d'etudiant a été annulée ")
            if ch2 != "" and ch3 != "":
                if not (existeCl(ch2,dataetd, 6)):
                    self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cette Section n'admet pas des "
                                                                                     "etudiants!")
                elif not (existeCl(ch3,dataetd, 6)):
                    self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING",
                                                             "Cette Niveau n'admet pas des etudiants!")
                elif not (existeDoubleCl(ch2,ch3,6,dataetd)):
                    self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING",
                                                             "Cette Classe n'admet pas des etudiants!")
                else:
                    self.box = QtWidgets.QMessageBox.question(self.window, "question",
                                                              "Voulez-vous vraiment enregistrer la suppression de la classe?")
                    if self.box == QMessageBox.Yes:
                        while (existeDoubleCl(ch2,ch3,6,dataetd)):
                            dataetd.remove(recherche2(dataetd, 6, ch2, ch3))

                        self.box = QtWidgets.QMessageBox.about(self.window, "Succés",
                                                               "La suppression de la classe est effectué avec succés")
                    else:
                        self.box = QtWidgets.QMessageBox.about(self.window, "Annuler",
                                                               "La suppression de la classe a été annulée ")

            elif ch2 != "" and ch3 == "":
                if not (existeCl(ch2, dataetd, 6)):
                    self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cette Section n'admet pas des "
                                                                                     "etudiants!")
                else:
                    self.box = QtWidgets.QMessageBox.question(self.window, "question",
                                                              "Voulez-vous vraiment enregistrer la suppression de Section?")
                    if self.box == QMessageBox.Yes:
                        while existeCl(ch2,dataetd, 6):
                            dataetd.remove(recherche2(dataetd, 6, ch2))

                        self.box = QtWidgets.QMessageBox.about(self.window, "Succés",
                                                               "La suppression de section est effectué avec succés")
                    else:
                        self.box = QtWidgets.QMessageBox.about(self.window, "Annuler",
                                                               "La suppression de Section a été annulée ")
            elif ch3 != "" and ch2 == "":
                if not (existeCl(ch3, dataetd, 6)):
                    self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cette Niveau n'admet pas des "
                                                                                     "etudiants!")
                else:
                    self.box = QtWidgets.QMessageBox.question(self.window, "question",
                                                              "Voulez-vous vraiment enregistrer la suppression de niveau?")
                    if self.box == QMessageBox.Yes:
                        while existeCl(ch3, dataetd, 6):
                            dataetd.remove(recherche2(dataetd, 6, ch2, ch3))

                        self.box = QtWidgets.QMessageBox.about(self.window, "Succés",
                                                               "La suppression de niveau est effectué avec succés")
                    else:
                        self.box = QtWidgets.QMessageBox.about(self.window, "Annuler",
                                                               "La suppression de niveau a été annulée ")
            enregistrement(dataetd, "etudiant.xlsx")


class modifierEtudiant:
    def __init__(self):
        self.window = QDialog()
        self.ui = Ui_Dialog2()
        self.ui.setupUi(self.window)
        self.ui.pushButton.clicked.connect(self.modifier)

    def modifier(self):
        ch = self.ui.lineEdit.text()
        ch2 = self.ui.lineEdit_2.text()
        ch3 = self.ui.lineEdit_3.text()
        ch4 = self.ui.lineEdit_4.text()
        if ch4 == "" or (ch2 == "" and ch3 == "" and ch == ""):
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Il n'y a pas assez d'informations")
        elif not (ch4.isnumeric() and len(ch4) == 8 and (ch == "" or (ch.isnumeric() and len(ch) == 8))):
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "N° d'inscription et N° téléphone "
                                                                             "doivent être 8 chiffres")
        elif ch2 != "" and (ch2[0] == "@" or ("@gmail.com" not in ch2)):
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Mail est incorrect")
        else:
            global dataetd
            if not (existe(ch4, dataetd, 2)):
                self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cet Etudiant n'existe pas!")
            else:
                if ch != "":
                    l=list(dataetd[dataetd.index(recherche(ch4,dataetd,2))])
                    l[5]=ch
                    dataetd[dataetd.index(recherche(ch4, dataetd, 2))]=l
                    self.box = QtWidgets.QMessageBox.about(self.window, "Succés","Le changement de N° de téléphone "
                                                                                 "est effectué avec succés")
                if ch2 != "":
                    l = list(dataetd[dataetd.index(recherche(ch4, dataetd, 2))])
                    l[3] = ch2
                    dataetd[dataetd.index(recherche(ch4, dataetd, 2))] = l
                    self.box = QtWidgets.QMessageBox.about(self.window, "Succés", "Le changement de Mail "
                                                                                  "est effectué avec succés")
                if ch3 != "":
                    l = list(dataetd[dataetd.index(recherche(ch4, dataetd, 2))])
                    l[4] = ch3
                    dataetd[dataetd.index(recherche(ch4, dataetd, 2))] = l
                    self.box = QtWidgets.QMessageBox.about(self.window, "Succés", "Le changement d'adresse' "
                                                                                  "est effectué avec succés")
                enregistrement(dataetd, "etudiant.xlsx")


class afficherEtudiant:
    def __init__(self):
        global sheet1
        self.window = QDialog()
        self.ui = Ui_Dialog10()
        self.ui.setupUi(self.window)
        self.ui.tableWidget.setRowCount(len(dataetd))
        self.ui.tableWidget.setColumnCount(sheet1.max_column)
        self.ui.tableWidget.setHorizontalHeaderLabels(dataetd[0])
        self.remplir_table()
        self.ui.pushButton.clicked.connect(self.tri)

    def remplir_table(self):
        global dataetd
        row = 0
        for v1 in dataetd[1:]:
            col = 0
            for v2 in v1:
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(v2)))
                col += 1
            row += 1
    def tri(self):
        global dataetd
        def take(elem):
            return elem[6]
        l=dataetd[1:]
        l.sort(key=take)
        row = 0
        for v1 in l:
            col = 0
            for v2 in v1:
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(v2)))
                col += 1
            row += 1


class recherche_par_id:
    def __init__(self):
        self.window = QDialog()
        self.ui = Ui_Dialog13()
        self.ui.setupUi(self.window)
        self.remplir_table()
        self.ui.pushButton.clicked.connect(self.recherche)
    def remplir_table(self):
        global dataetd
        self.ui.tableWidget.setRowCount(sheet1.max_row)
        self.ui.tableWidget.setColumnCount(sheet1.max_column)
        self.ui.tableWidget.setHorizontalHeaderLabels(dataetd[0])
        row = 0
        for v1 in dataetd[1:]:
            col = 0
            for v2 in v1:
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(v2)))
                col += 1
            row += 1
    def recherche(self):
        global dataetd
        ch=self.ui.lineEdit.text()
        if ch=="":
            self.remplir_table()
        else:
            if not(existe(ch,dataetd,2)):
                self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cet Etudiant n'existe pas!")
            else:
                self.ui.tableWidget.setRowCount(1)
                self.ui.tableWidget.setColumnCount(sheet1.max_column)
                self.ui.tableWidget.setHorizontalHeaderLabels(dataetd[0])
                c=0
                for i in recherche(ch,dataetd,2):
                    self.ui.tableWidget.setItem(0,c,QTableWidgetItem(str(i)))
                    c+=1


class recherche_par_sec:
    def __init__(self):
        global dataetd
        global sheet1
        self.window = QDialog()
        self.ui = Ui_Dialog14()
        self.ui.setupUi(self.window)
        self.remplir_table()
        self.ui.pushButton.clicked.connect(self.recherche)
    def remplir_table(self):
        self.ui.tableWidget.setRowCount(sheet1.max_row)
        self.ui.tableWidget.setColumnCount(sheet1.max_column)
        self.ui.tableWidget.setHorizontalHeaderLabels(dataetd[0])
        row = 0
        for v1 in dataetd[1:]:
            col = 0
            for v2 in v1:
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(v2)))
                col += 1
            row += 1
    def recherche(self):
        ch=self.ui.comboBox.currentText()
        if ch=="":
            self.remplir_table()
        else:
            if not(existeCl(ch, dataetd, 6)):
                self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cette Section n'admet pas des "
                                                                                     "etudiants!")
            else:
                self.ui.tableWidget.setRowCount(sheet1.max_row)
                self.ui.tableWidget.setColumnCount(sheet1.max_column)
                self.ui.tableWidget.setHorizontalHeaderLabels(dataetd[0])
                r=0
                l=dataetd
                for v1 in dataetd[1:]:
                    c = 0
                    if existeCl1(ch,v1,6):
                        for i in v1:
                            self.ui.tableWidget.setItem(r, c, QTableWidgetItem(str(i)))
                            c += 1
                        r+=1
                self.ui.tableWidget.setRowCount(r)


class recherche_par_niv:
    def __init__(self):
        global dataetd
        global sheet1
        self.window = QDialog()
        self.ui = Ui_Dialog15()
        self.ui.setupUi(self.window)
        self.remplir_table()
        self.ui.pushButton.clicked.connect(self.recherche)
    def remplir_table(self):
        self.ui.tableWidget.setRowCount(sheet1.max_row)
        self.ui.tableWidget.setColumnCount(sheet1.max_column)
        self.ui.tableWidget.setHorizontalHeaderLabels(dataetd[0])
        row = 0
        for v1 in dataetd[1:]:
            col = 0
            for v2 in v1:
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(v2)))
                col += 1
            row += 1
    def recherche(self):
        ch=self.ui.comboBox_2.currentText()
        if ch=="":
            self.remplir_table()
        else:
            if not(existeCl(ch, dataetd, 6)):
                self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cet niveau n'admet pas des "
                                                                                     "etudiants!")
            else:
                self.ui.tableWidget.setRowCount(sheet1.max_row)
                self.ui.tableWidget.setColumnCount(sheet1.max_column)
                self.ui.tableWidget.setHorizontalHeaderLabels(dataetd[0])
                r=0
                for v1 in dataetd[1:]:
                    col = 0

                    if ch in v1[6]:
                        for v2 in v1:
                            self.ui.tableWidget.setItem(r, col, QTableWidgetItem(str(v2)))
                            col += 1
                        r += 1
                self.ui.tableWidget.setRowCount(r)


class recherche_par_cl:
    def __init__(self):
        self.window = QDialog()
        self.ui = Ui_Dialog16()
        self.ui.setupUi(self.window)
        self.remplir_table()
        self.ui.pushButton.clicked.connect(self.recherche)
    def remplir_table(self):
        global dataetd
        global sheet1
        self.ui.tableWidget.setRowCount(sheet1.max_row)
        self.ui.tableWidget.setColumnCount(sheet1.max_column)
        self.ui.tableWidget.setHorizontalHeaderLabels(dataetd[0])
        row = 0
        for v1 in dataetd[1:]:
            col = 0
            for v2 in v1:
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(v2)))
                col += 1
            row += 1
    def recherche(self):
        global dataetd
        global sheet1
        ch=self.ui.comboBox.currentText()
        if ch=="":
            self.remplir_table()
        else:
            if not(existe(ch,dataetd,6)):
                self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cette classe n'admet pas des "
                                                                                     "etudiants!")
            else:
                self.ui.tableWidget.setRowCount(sheet1.max_row)
                self.ui.tableWidget.setColumnCount(sheet1.max_column)
                self.ui.tableWidget.setHorizontalHeaderLabels(dataetd[0])
                r=0
                for v1 in dataetd[1:]:
                    c = 0
                    if ch==v1[6]:
                        for i in v1:
                            self.ui.tableWidget.setItem(r, c, QTableWidgetItem(str(i)))
                            c += 1
                        r+=1
                self.ui.tableWidget.setRowCount(r)


class ajouterLivre:
    def __init__(self):
        self.window = QDialog()
        self.ui = Ui_Dialog3()
        self.ui.setupUi(self.window)
        self.ui.pushButton.clicked.connect(self.ajouter)

    def ajouter(self):
        global datalv
        ch1 = self.ui.lineEdit.text()
        ch2 = self.ui.lineEdit_2.text()
        ch3 = self.ui.lineEdit_3.text()
        ch4 = self.ui.lineEdit_4.text()
        ch5 = self.ui.spinBox.text()
        ch6 = self.ui.dateEdit.text()
        if ch5 == "":
            ch5 = "0"
        if ch1 == "" or ch2 == "" or ch3 == "" or ch4 == "" or ch6 == "":
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Il y a un champ vide")
        else:
            if ch5 == "0":
                self.box = QtWidgets.QMessageBox.question(self.window, "question", "Voulez-vous vraiment enregistrer "
                                                                                   "le nombre de livre est égale à "
                                                                                   "'0'?")
            if ch5 != "0" or self.box == QMessageBox.Yes:
                if existe(ch1, datalv, 0):
                    self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cet livre déja existe!")
                else:
                    ch3 = ch3 + " " + ch4
                    l = [ch1, ch2, majus(ch3), ch6, ch5]
                    datalv.append(l)
                    enregistrement(l, "livre.xlsx")
                    self.box = QtWidgets.QMessageBox.about(self.window, "Succés", "L'ajout est effectué avec succés")


class supprimerLivre:
    def __init__(self):
        self.window = QDialog()
        self.ui = Ui_Dialog4()
        self.ui.setupUi(self.window)
        self.ui.pushButton.clicked.connect(self.supprimer)

    def supprimer(self):
        ch1 = self.ui.lineEdit.text()
        ch2 = self.ui.lineEdit_2.text()
        ch3 = self.ui.lineEdit_3.text()
        if ch1 == "" and ch2 == "" and ch3 == "":
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "il faut remplir au moins un champ")
        if not (ch1 == "" and ch2 == "" and ch3 == ""):
            global datalv
            if ch1 != "":
                if not (existe(ch1, datalv, 0)):
                    self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cet Livre n'existe pas!")
                else:
                    self.box = QtWidgets.QMessageBox.question(self.window, "question",
                                                              "Voulez-vous vraiment enregistrer la suppression du Livre?")
                    if self.box == QMessageBox.Yes:
                        datalv.remove(recherche(ch1, datalv, 0))

                        self.box = QtWidgets.QMessageBox.about(self.window, "Succés",
                                                               "La suppression du Livre est effectué "
                                                               "avec succés")
                    else:
                        self.box = QtWidgets.QMessageBox.about(self.window, "Annuler",
                                                               "La suppression du Livre a été annulée ")
            if ch2 != "":
                if not (existe(majus(ch2), datalv, 2)):
                    self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cet Auteur n'existe pas!")
                else:
                    self.box=QtWidgets.QMessageBox.question(self.window, "question", "Voulez-vous vraiment "
                                                                                       "enregistrer la suppression des"
                                                                                       " Livres de '" + ch2 + "'?")
                    if self.box == QMessageBox.Yes:
                        while existe(majus(ch2), datalv, 2):
                            datalv.remove(recherche(majus(ch2), datalv, 2))

                        self.box = QtWidgets.QMessageBox.about(self.window, "Succés",
                                                               "La suppression des Livres de '" + ch2 + "' est effectué avec succés")
                    else:
                        self.box = QtWidgets.QMessageBox.about(self.window, "Annuler",
                                                               "La suppression des Livres par auteur a été annulée ")

            if ch3 != "":
                self.box = QtWidgets.QMessageBox.question(self.window, "question", "Voulez-vous vraiment "
                                                                                   "enregistrer la suppression des"
                                                                                   " Livres ont édité dans " + ch3 + "?")
                if self.box == QMessageBox.Yes:
                    while existe(ch3, datalv, 3):
                        datalv.remove(recherche(ch3, datalv, 3))

                    self.box = QtWidgets.QMessageBox.about(self.window, "Succés",
                                                           "La suppression des Livres de " + ch2 + " est effectué avec succés")
                else:
                    self.box = QtWidgets.QMessageBox.about(self.window, "Annuler",
                                                           "La suppression des Livres par année a été annulée ")
        enregistrement(datalv, "livre.xlsx")


class modifierLivre:
    def __init__(self):
        self.window = QDialog()
        self.ui = Ui_Dialog5()
        self.ui.setupUi(self.window)
        self.ui.pushButton.clicked.connect(self.modifier)

    def modifier(self):
        ch1 = self.ui.lineEdit_2.text()
        ch2 = self.ui.spinBox.text()
        if ch2 == "":
            ch2 = "0"
        if ch1 == "":
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Il n'y a pas assez d'informations")
        else:
            if ch2 == "0":
                self.box = QtWidgets.QMessageBox.question(self.window, "question", "Voulez-vous vraiment enregistrer "
                                                                                   "le nombre de livre est égale à "
                                                                                   "'0'?")
            if ch2 != "0" or self.box == QMessageBox.Yes:
                global datalv
                if not (existe(ch1, datalv, 0)):
                    self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cet Livre n'existe pas!")
                else:

                    l = list(datalv[datalv.index(recherche(ch1, datalv, 0))])
                    l[4] = ch2
                    datalv[datalv.index(recherche(ch1, datalv, 0))] = l
                    self.box = QtWidgets.QMessageBox.about(self.window, "Succés", "Le changement du Nombre "
                                                                                  "est effectué avec succés")
                    enregistrement(datalv, "livre.xlsx")


class afficherLivre:
    def __init__(self):
        global sheet2
        global datalv
        self.window = QDialog()
        self.ui = Ui_Dialog11()
        self.ui.setupUi(self.window)
        self.ui.tableWidget.setRowCount(sheet2.max_row)
        self.ui.tableWidget.setColumnCount(sheet2.max_column)
        self.ui.tableWidget.setHorizontalHeaderLabels(datalv[0])
        self.remplir_table()
        self.ui.pushButton.clicked.connect(self.tri)

    def remplir_table(self):
        row = 0
        for v1 in datalv[1:]:
            col = 0
            for v2 in v1:
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(v2)))
                col += 1
            row += 1

    def tri(self):
        def take(elem):
            return elem[1]

        l = datalv[1:]
        l.sort(key=take)
        row = 0
        for v1 in l:
            col = 0
            for v2 in v1:
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(v2)))
                col += 1
            row += 1


class recherche_par_ref:
    def __init__(self):
        global sheet2
        global datalv
        self.window = QDialog()
        self.ui = Ui_Dialog17()
        self.ui.setupUi(self.window)
        self.remplir_table()
        self.ui.pushButton.clicked.connect(self.recherche)
    def remplir_table(self):
        self.ui.tableWidget.setRowCount(sheet2.max_row)
        self.ui.tableWidget.setColumnCount(sheet2.max_column)
        self.ui.tableWidget.setHorizontalHeaderLabels(datalv[0])
        row = 0
        for v1 in datalv[1:]:
            col = 0
            for v2 in v1:
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(v2)))
                col += 1
            row += 1
    def recherche(self):
        ch=self.ui.lineEdit.text()
        if ch=="":
            self.remplir_table()
        else:
            if not(existe(ch,datalv,0)):
                self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cet Livre n'existe pas!")
            else:
                self.ui.tableWidget.setRowCount(1)
                self.ui.tableWidget.setColumnCount(sheet2.max_column)
                self.ui.tableWidget.setHorizontalHeaderLabels(datalv[0])
                c=0
                for i in recherche(ch,datalv,0):
                    self.ui.tableWidget.setItem(0,c,QTableWidgetItem(str(i)))
                    c+=1


class recherche_par_tit:
    def __init__(self):
        global sheet2
        global datalv
        self.window = QDialog()
        self.ui = Ui_Dialog18()
        self.ui.setupUi(self.window)
        self.remplir_table()
        self.ui.pushButton.clicked.connect(self.recherche)
    def remplir_table(self):
        self.ui.tableWidget.setRowCount(sheet2.max_row)
        self.ui.tableWidget.setColumnCount(sheet2.max_column)
        self.ui.tableWidget.setHorizontalHeaderLabels(datalv[0])
        row = 0
        for v1 in datalv[1:]:
            col = 0
            for v2 in v1:
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(v2)))
                col += 1
            row += 1
    def recherche(self):
        ch=self.ui.lineEdit.text()
        if ch=="":
            self.remplir_table()
        else:
            if not(existeIn(ch, datalv, 1)):
                self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cet titre n'existe pas!")
            else:
                self.ui.tableWidget.setRowCount(sheet2.max_row)
                self.ui.tableWidget.setColumnCount(sheet2.max_column)
                self.ui.tableWidget.setHorizontalHeaderLabels(datalv[0])
                r = 0
                for v1 in datalv[1:]:
                    c = 0
                    if ch in v1[1]:
                        for i in v1:
                            self.ui.tableWidget.setItem(r, c, QTableWidgetItem(str(i)))
                            c += 1
                        r += 1
                self.ui.tableWidget.setRowCount(r)


class recherche_par_aut:
    def __init__(self):
        self.window = QDialog()
        self.ui = Ui_Dialog19()
        self.ui.setupUi(self.window)
        self.remplir_table()
        self.ui.pushButton.clicked.connect(self.recherche)
    def remplir_table(self):
        self.ui.tableWidget.setRowCount(sheet2.max_row)
        self.ui.tableWidget.setColumnCount(sheet2.max_column)
        self.ui.tableWidget.setHorizontalHeaderLabels(datalv[0])
        row = 0
        for v1 in datalv[1:]:
            col = 0
            for v2 in v1:
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(v2)))
                col += 1
            row += 1
    def recherche(self):
        global datalv
        global sheet2
        ch=self.ui.lineEdit.text()
        if ch=="":
            self.remplir_table()
        else:
            ch=majus(ch)
            if not(existeIn(ch, datalv, 2)):
                print("hi")
                self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cet auteur n'existe pas!")
            else:
                self.ui.tableWidget.setRowCount(sheet2.max_row)
                self.ui.tableWidget.setColumnCount(sheet2.max_column)
                self.ui.tableWidget.setHorizontalHeaderLabels(datalv[0])
                r = 0
                for v1 in datalv[1:]:
                    c = 0
                    if ch in v1[2]:
                        for i in v1:
                            self.ui.tableWidget.setItem(r, c, QTableWidgetItem(str(i)))
                            c += 1
                        r += 1
                self.ui.tableWidget.setRowCount(r)


class recherche_par_ann:
    def __init__(self):
        self.window = QDialog()
        self.ui = Ui_Dialog20()
        self.ui.setupUi(self.window)
        self.remplir_table()
        self.ui.pushButton.clicked.connect(self.recherche)
    def remplir_table(self):
        self.ui.tableWidget.setRowCount(sheet2.max_row)
        self.ui.tableWidget.setColumnCount(sheet2.max_column)
        self.ui.tableWidget.setHorizontalHeaderLabels(datalv[0])
        row = 0
        for v1 in datalv[1:]:
            col = 0
            for v2 in v1:
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(v2)))
                col += 1
            row += 1
    def recherche(self):
        global datalv
        global sheet2
        ch=self.ui.lineEdit.text()
        if ch=="":
            self.remplir_table()
        else:
            if not(existe(ch,datalv,3)):
                self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cet Année n'existe pas!")
            else:
                self.ui.tableWidget.setRowCount(sheet2.max_row)
                self.ui.tableWidget.setColumnCount(sheet2.max_column)
                self.ui.tableWidget.setHorizontalHeaderLabels(datalv[0])
                r = 0
                for v1 in datalv[1:]:
                    c = 0
                    if v1[3]==ch:
                        for i in v1:
                            self.ui.tableWidget.setItem(r, c, QTableWidgetItem(str(i)))
                            c += 1
                        r += 1
                self.ui.tableWidget.setRowCount(r)


class ajouterEmprunt:
    def __init__(self):
        self.window = QDialog()
        self.ui = Ui_Dialog6()
        self.ui.setupUi(self.window)
        self.ui.pushButton.clicked.connect(self.ajouter)

    def ajouter(self):
        ch1 = self.ui.lineEdit.text()
        ch2 = self.ui.lineEdit_2.text()
        ch3 = str(date.today())
        ch3=ch3[:4]+"/"+ch3[5:7]+"/"+ch3[8:]
        ch3="/".join(ch3.split("/")[::-1])
        if ch1 == "" or ch2 == "":
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Il y a un champ vide")
        elif not (ch1.isnumeric()) or len(ch1) != 8:
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "N° d'inscription doit être 8 chiffres")
        else:
            global dataetd
            global datalv
            global dataemp
            if not (existe(ch1, dataetd, 2)):
                self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cette etudiant n'existe pas!")
            elif not (existe(ch2, datalv, 0)):
                self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cette livre n'existe pas!")
            else:
                if int(recherche(ch2,datalv,0)[4])==0:
                    self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "le nombre de livre est insuffisant!")
                else:
                    l=list(recherche(ch2, datalv, 0))
                    l[4]=str(int(l[4])-1)
                    datalv[datalv.index(recherche(ch2, datalv, 0))]=l
                    l = [ch1, ch2, ch3,"None"]
                    dataemp.append(l)
                    enregistrement(l, "emprunt.xlsx")
                    enregistrement(datalv, "livre.xlsx")
                    self.box = QtWidgets.QMessageBox.about(self.window, "Succés", "L'ajout est effectué avec succés")


class supprimerEmprunt:
    def __init__(self):
        self.window = QDialog()
        self.ui = Ui_Dialog7()
        self.ui.setupUi(self.window)
        self.ui.pushButton.clicked.connect(self.supprimer)

    def supprimer(self):
        ch1 = self.ui.lineEdit.text()
        ch2 = self.ui.lineEdit_2.text()
        if ch1 == "" or ch2 == "":
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Il y a un champ vide")
        elif not (ch1.isnumeric() and len(ch1) == 8):
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "N° d'inscription doit être 8 chiffres")
        else:
            global dataemp
            if not (existeDouble(ch1,0,ch2,1,dataemp)):
                self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING","Cet emprunt n'existe pas")
            else:
                self.box = QtWidgets.QMessageBox.question(self.window, "question",
                                                          "Voulez-vous vraiment enregistrer la suppression d'emprunt?")
                if self.box == QMessageBox.Yes:
                    dataemp.remove(recherche3(ch1,0,ch2,1,dataemp))
                    enregistrement(dataemp, "emprunt.xlsx")
                    self.box = QtWidgets.QMessageBox.about(self.window, "Succés",
                                                           "La suppression d'emprunt' est effectué "
                                                           "avec succés")
                else:
                    self.box = QtWidgets.QMessageBox.about(self.window, "Annuler",
                                                           "La suppression d'emprunt a été annulée ")


class modifierDateEmp:
    def __init__(self):
        self.window = QDialog()
        self.ui = Ui_Dialog8()
        self.ui.setupUi(self.window)
        self.ui.pushButton.clicked.connect(self.modifierEmp)

    def modifierEmp(self):
        ch1 = self.ui.lineEdit.text()
        ch2 = self.ui.lineEdit_2.text()
        ch3 = self.ui.dateEdit.text()
        if ch1 == "" or ch2 == "":
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Il y a un champ vide")
        elif not (ch1.isnumeric() and len(ch1) == 8):
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "N° d'inscription doit être 8 chiffres")
        else:
            global dataemp
            if not (existeDouble(ch1, 0, ch2, 1, dataemp)):
                self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cet emprunt n'existe pas")
            else:
                self.box = QtWidgets.QMessageBox.question(self.window, "question",
                                                          "Voulez-vous vraiment enregistrer la modification d'emprunt?")
                if self.box == QMessageBox.Yes:
                    l = list(dataemp[dataemp.index(recherche3(ch1, 0, ch2, 1, dataemp))])
                    if l[3]!="None" and dateinf(l[3],ch3):
                        self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING",
                                                                 "La date de retour ne peut pas être "
                                                                 "inférieur à la date d'emprunt!")
                    else:
                        l[2] = ch3
                        dataemp[dataemp.index(recherche3(ch1, 0, ch2, 1, dataemp))] = l
                        self.box = QtWidgets.QMessageBox.about(self.window, "Succés", "Le changement du Date d'emprunt "
                                                                                      "est effectué avec succés")
                        enregistrement(dataemp, "emprunt.xlsx")


class modifierDateRet:
    def __init__(self):
        self.window = QDialog()
        self.ui = Ui_Dialog9()
        self.ui.setupUi(self.window)
        self.ui.pushButton.clicked.connect(self.modifierRet)

    def modifierRet(self):
        ch1 = self.ui.lineEdit.text()
        ch2 = self.ui.lineEdit_2.text()
        ch3 = self.ui.dateEdit.text()
        if ch1 == "" or ch2 == "":
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Il y a un champ vide")
        elif not (ch1.isnumeric() and len(ch1) == 8):
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "N° d'inscription doit être 8 chiffres")
        else:
            global dataemp
            if not (existeDouble(ch1, 0, ch2, 1, dataemp)):
                self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cet emprunt n'existe pas")
            else:
                self.box = QtWidgets.QMessageBox.question(self.window, "question",
                                                          "Voulez-vous vraiment enregistrer la modification d'emprunt?")
                if self.box == QMessageBox.Yes:
                    l = list(dataemp[dataemp.index(recherche3(ch1,0,ch2,1,dataemp))])
                    if l[3]=="None":
                        self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING",
                                                                 "La date de retour de cette Emprunt n'existe pas ")
                    elif dateinf(ch3,l[2]):
                        self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING",
                                                                 "La date de retour ne peut pas être "
                                                                 "inférieur à la date d'emprunt!")
                    else:
                        l[3] = ch3
                        dataemp[dataemp.index(recherche3(ch1,0,ch2,1,dataemp))] = l
                        self.box = QtWidgets.QMessageBox.about(self.window, "Succés", "Le changement du Date d'emprunt "
                                                                                      "est effectué avec succés")
                        enregistrement(dataemp, "emprunt.xlsx")


class afficherEmprunt:
    def __init__(self):
        global dataemp
        global sheet3
        self.window = QDialog()
        self.ui = Ui_Dialog12()
        self.ui.setupUi(self.window)
        self.ui.tableWidget.setRowCount(sheet3.max_row)
        self.ui.tableWidget.setColumnCount(sheet3.max_column)
        self.ui.tableWidget.setHorizontalHeaderLabels(dataemp[0])
        self.remplir_table()

    def remplir_table(self):
        row = 0
        for v1 in dataemp[1:]:
            col = 0
            for v2 in v1:
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(v2)))
                col += 1
            row += 1


class recherche_par_etd:
    def __init__(self):
        global dataemp
        global sheet3
        self.window = QDialog()
        self.ui = Ui_Dialog21()
        self.ui.setupUi(self.window)
        self.remplir_table()
        self.ui.pushButton.clicked.connect(self.recherche)
    def remplir_table(self):
        self.ui.tableWidget.setRowCount(sheet3.max_row)
        self.ui.tableWidget.setColumnCount(sheet3.max_column)
        self.ui.tableWidget.setHorizontalHeaderLabels(dataemp[0])
        row = 0
        for v1 in dataemp[1:]:
            col = 0
            for v2 in v1:
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(v2)))
                col += 1
            row += 1
    def recherche(self):
        ch=self.ui.lineEdit.text()
        if ch=="":
            self.remplir_table()
        else:
            if not(existe(ch,dataemp,0)):
                self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cet étudiant n'admet pas des emprunt!")
            else:
                self.ui.tableWidget.setRowCount(sheet3.max_row)
                self.ui.tableWidget.setColumnCount(sheet3.max_column)
                self.ui.tableWidget.setHorizontalHeaderLabels(dataemp[0])
                r = 0
                l = dataemp
                while existe(ch, l, 0):
                    c = 0
                    for i in recherche(ch, l, 0):
                        self.ui.tableWidget.setItem(r, c, QTableWidgetItem(str(i)))
                        c += 1
                    r += 1
                    l.remove(recherche(ch, l, 0))
                self.ui.tableWidget.setRowCount(r)


class recherche_par_liv:
    def __init__(self):
        self.window = QDialog()
        self.ui = Ui_Dialog22()
        self.ui.setupUi(self.window)
        self.remplir_table()
        self.ui.pushButton.clicked.connect(self.recherche)
    def remplir_table(self):
        self.ui.tableWidget.setRowCount(sheet3.max_row)
        self.ui.tableWidget.setColumnCount(sheet3.max_column)
        self.ui.tableWidget.setHorizontalHeaderLabels(dataemp[0])
        row = 0
        for v1 in dataemp[1:]:
            col = 0
            for v2 in v1:
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(v2)))
                col += 1
            row += 1
    def recherche(self):
        ch=self.ui.lineEdit.text()
        if ch=="":
            self.remplir_table()
        else:
            if not(existe(ch,dataemp,1)):
                self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cet livre n'admet pas des emprunt!")
            else:
                self.ui.tableWidget.setRowCount(sheet3.max_row)
                self.ui.tableWidget.setColumnCount(sheet3.max_column)
                self.ui.tableWidget.setHorizontalHeaderLabels(dataemp[0])
                r = 0
                l = dataemp
                while existe(ch, l, 1):
                    c = 0
                    for i in recherche(ch, l, 1):
                        self.ui.tableWidget.setItem(r, c, QTableWidgetItem(str(i)))
                        c += 1
                    r += 1
                    l.remove(recherche(ch, l, 1))
                self.ui.tableWidget.setRowCount(r)


class recherche_par_demp:
    def __init__(self):
        self.window = QDialog()
        self.ui = Ui_Dialog23()
        self.ui.setupUi(self.window)
        self.remplir_table()
        self.ui.pushButton.clicked.connect(self.recherche)
    def remplir_table(self):
        self.workbook = openpyxl.load_workbook("emprunt.xlsx")
        self.sheet = self.workbook.active
        self.data = list(self.sheet.values)
        self.ui.tableWidget.setRowCount(self.sheet.max_row)
        self.ui.tableWidget.setColumnCount(self.sheet.max_column)
        self.ui.tableWidget.setHorizontalHeaderLabels(self.data[0])
        row = 0
        for v1 in self.data[1:]:
            col = 0
            for v2 in v1:
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(v2)))
                col += 1
            row += 1
    def recherche(self):
        ch=self.ui.dateEdit.text()
        if ch=="":
            self.remplir_table()
        else:
            self.workbook = openpyxl.load_workbook("emprunt.xlsx")
            self.sheet = self.workbook.active
            self.data = list(self.sheet.values)
            if not(existe(ch,self.data,2)):
                self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cet date d'emprunt n'existe pas!")
            else:
                self.ui.tableWidget.setRowCount(self.sheet.max_row)
                self.ui.tableWidget.setColumnCount(self.sheet.max_column)
                self.ui.tableWidget.setHorizontalHeaderLabels(self.data[0])
                r = 0
                l = self.data
                while existe(ch, l, 2):
                    c = 0
                    for i in recherche(ch, l, 2):
                        self.ui.tableWidget.setItem(r, c, QTableWidgetItem(str(i)))
                        c += 1
                    r += 1
                    l.remove(recherche(ch, l, 2))
                self.ui.tableWidget.setRowCount(r)


class recherche_par_dret:
    def __init__(self):
        global dataemp
        global sheet3
        self.window = QDialog()
        self.ui = Ui_Dialog24()
        self.ui.setupUi(self.window)
        self.remplir_table()
        self.ui.pushButton.clicked.connect(self.recherche)
    def remplir_table(self):
        self.ui.tableWidget.setRowCount(sheet3.max_row)
        self.ui.tableWidget.setColumnCount(sheet3.max_column)
        self.ui.tableWidget.setHorizontalHeaderLabels(dataemp[0])
        row = 0
        for v1 in dataemp[1:]:
            col = 0
            for v2 in v1:
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(v2)))
                col += 1
            row += 1
    def recherche(self):
        ch=self.ui.dateEdit.text()
        if ch=="":
            self.remplir_table()
        else:
            if not(existe(ch,dataemp,3)):
                self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cet date d'emprunt n'existe pas!")
            else:
                self.ui.tableWidget.setRowCount(sheet3.max_row)
                self.ui.tableWidget.setColumnCount(sheet3.max_column)
                self.ui.tableWidget.setHorizontalHeaderLabels(dataemp[0])
                r = 0
                l = dataemp
                while existe(ch, l, 3):
                    c = 0
                    for i in recherche(ch, l, 3):
                        self.ui.tableWidget.setItem(r, c, QTableWidgetItem(str(i)))
                        c += 1
                    r += 1
                    l.remove(recherche(ch, l, 3))
                self.ui.tableWidget.setRowCount(r)


class recherche_entre_demp:
    def __init__(self):
        global dataemp
        global sheet3
        self.window = QDialog()
        self.ui = Ui_Dialog25()
        self.ui.setupUi(self.window)
        self.remplir_table()
        self.ui.pushButton.clicked.connect(self.recherche)
    def remplir_table(self):
        self.ui.tableWidget.setRowCount(sheet3.max_row)
        self.ui.tableWidget.setColumnCount(sheet3.max_column)
        self.ui.tableWidget.setHorizontalHeaderLabels(dataemp[0])
        row = 0
        for v1 in dataemp[1:]:
            col = 0
            for v2 in v1:
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(v2)))
                col += 1
            row += 1
    def recherche(self):
        ch1=self.ui.dateEdit.text()
        ch2=self.ui.dateEdit_2.text()
        self.ui.tableWidget.setRowCount(sheet3.max_row)
        self.ui.tableWidget.setColumnCount(sheet3.max_column)
        self.ui.tableWidget.setHorizontalHeaderLabels(dataemp[0])
        r = 0
        for j in dataemp[1:]:
            c = 0
            if dateinf(ch1,j[2]) and dateinf(j[2],ch2):
                for i in j:
                    self.ui.tableWidget.setItem(r, c, QTableWidgetItem(str(i)))
                    c += 1
                r += 1
        self.ui.tableWidget.setRowCount(r)


class recherche_entre_dret:
    def __init__(self):
        self.window = QDialog()
        self.ui = Ui_Dialog26()
        self.ui.setupUi(self.window)
        self.remplir_table()
        self.ui.pushButton.clicked.connect(self.recherche)
    def remplir_table(self):
        self.ui.tableWidget.setRowCount(sheet3.max_row)
        self.ui.tableWidget.setColumnCount(sheet3.max_column)
        self.ui.tableWidget.setHorizontalHeaderLabels(dataemp[0])
        row = 0
        for v1 in dataemp[1:]:
            col = 0
            for v2 in v1:
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(v2)))
                col += 1
            row += 1
    def recherche(self):
        global dataemp
        global sheet3
        ch1=self.ui.dateEdit.text()
        ch2=self.ui.dateEdit_2.text()
        self.ui.tableWidget.setRowCount(sheet3.max_row)
        self.ui.tableWidget.setColumnCount(sheet3.max_column)
        self.ui.tableWidget.setHorizontalHeaderLabels(dataemp[0])
        r = 0
        for j in dataemp[1:]:
            c = 0
            if j[3]!="None" and dateinf(ch1,j[3]) and dateinf(j[3],ch2):
                for i in j:
                    self.ui.tableWidget.setItem(r, c, QTableWidgetItem(str(i)))
                    c += 1
                r += 1
        self.ui.tableWidget.setRowCount(r)


class retourEmp:
    def __init__(self):
        self.window = QDialog()
        self.ui = Ui_Dialog27()
        self.ui.setupUi(self.window)
        self.ui.pushButton.clicked.connect(self.modifierEmp)

    def modifierEmp(self):
        global datalv
        ch1 = self.ui.lineEdit.text()
        ch2 = self.ui.lineEdit_2.text()
        ch3 = str(date.today())
        ch3 = ch3[:4] + "/" + ch3[5:7] + "/" + ch3[8:]
        ch3="/".join(ch3.split("/")[::-1])
        if ch1 == "" or ch2 == "":
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Il y a un champ vide")
        elif not (ch1.isnumeric() and len(ch1) == 8):
            self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "N° d'inscription doit être 8 chiffres")
        else:
            global dataemp
            if not (existeDouble(ch1, 0, ch2, 1, dataemp)):
                self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING", "Cet emprunt n'existe pas")
            else:
                self.box = QtWidgets.QMessageBox.question(self.window, "question",
                                                          "Voulez-vous vraiment enregistrer la modification d'emprunt?")
                if self.box == QMessageBox.Yes:
                    l = list(dataemp[dataemp.index(recherche3(ch1, 0, ch2, 1, dataemp))])
                    if l[3]!="None" :
                        self.box = QtWidgets.QMessageBox.warning(self.window, "WARNING",
                                                                 "l'Emprunt est déja retourné ")
                    else:
                        l[3] = ch3
                        dataemp[dataemp.index(recherche3(ch1, 0, ch2, 1, dataemp))] = l
                        l = list(recherche(ch2, datalv, 0))
                        l[4] = str(int(l[4]) + 1)
                        datalv[datalv.index(recherche(ch2, datalv, 0))] = l
                        self.box = QtWidgets.QMessageBox.about(self.window, "Succés", "Le retour d'emprunt "
                                                                                      "est effectué avec succés")
                        enregistrement(dataemp, "emprunt.xlsx")
                        enregistrement(datalv, "livre.xlsx")